import os
import re
import json
import time
import uuid
import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from google import genai
from http.server import HTTPServer, BaseHTTPRequestHandler
from threading import Thread
import PIL.Image
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TELEGRAM_TOKEN = os.environ['TELEGRAM_TOKEN']
GEMINI_KEY = os.environ['GEMINI_KEY']

DEFAULT_CAPITAL = 20000
DEFAULT_RISK_PERCENT = 1
DEFAULT_LEVERAGE = 100
DEFAULT_CURRENCY = 'GBP'

JOURNAL_FILE = 'journal.json'

client = genai.Client(api_key=GEMINI_KEY)
bot = telebot.TeleBot(TELEGRAM_TOKEN)

user_settings = {}
pending_trades = {}
trade_counter = 0
update_states = {}

# ── Journal helpers ──────────────────────────────────────────────────────────

def load_journal():
    if os.path.exists(JOURNAL_FILE):
        with open(JOURNAL_FILE, 'r') as f:
            return json.load(f)
    return []

def save_journal(entries):
    with open(JOURNAL_FILE, 'w') as f:
        json.dump(entries, f, indent=2)

def find_entry_by_uid(entries, uid):
    for i, e in enumerate(entries):
        if e.get('uid') == uid:
            return i, e
    return None, None

def parse_analysis(text):
    fields = {}
    patterns = {
        'pair':     r'Pair:\s*(.+)',
        'entry':    r'Entry:\s*(.+)',
        'sl':       r'SL:\s*(.+)',
        'distance': r'Distance:\s*(.+)',
        'lot':      r'LOT:\s*(.+)',
        'tp2':      r'RR 1:2.*?TP:\s*(.+)',
        'tp3':      r'RR 1:3.*?TP:\s*(.+)',
        'margin':   r'Margin:\s*(.+)',
    }
    clean = re.sub(r'<[^>]+>', '', text)
    for key, pattern in patterns.items():
        m = re.search(pattern, clean)
        fields[key] = m.group(1).strip() if m else ''
    return fields

# ── Settings helpers ─────────────────────────────────────────────────────────

def get_settings(chat_id):
    if chat_id not in user_settings:
        user_settings[chat_id] = {
            'capital': DEFAULT_CAPITAL,
            'risk': DEFAULT_RISK_PERCENT,
            'leverage': DEFAULT_LEVERAGE,
            'currency': DEFAULT_CURRENCY,
        }
    s = user_settings[chat_id]
    if 'currency' not in s:
        s['currency'] = DEFAULT_CURRENCY
    return s

# ── Commands ─────────────────────────────────────────────────────────────────

@bot.message_handler(commands=['start'])
def handle_start(message):
    chat_id = message.chat.id
    s = get_settings(chat_id)
    cur = s['currency']
    risk_amt = s['capital'] * s['risk'] / 100
    text = (
        f"<b>Welcome to FX Position Size Calculator 📊</b>\n\n"
        f"<b>Your current settings:</b>\n"
        f"💰 Account size: <b>{cur} {s['capital']:,}</b>\n"
        f"⚡ Leverage: <b>1:{s['leverage']}</b>\n"
        f"🎯 Risk per trade: <b>{s['risk']}%</b> ({cur} {risk_amt:,.2f})\n"
        f"💱 Account currency: <b>{cur}</b>\n\n"
        f"<b>Commands:</b>\n"
        f"/capital [amount] — set account size\n"
        f"/risk [percent] — set risk per trade\n"
        f"/leverage [value] — set leverage\n"
        f"/currency [code] — set account currency (e.g. GBP, USD, EUR)\n"
        f"/settings — show current settings\n"
        f"/journal — view your trading journal\n"
        f"/export — export today's trades to Excel\n"
        f"/export [date] — export specific date (e.g. /export 2026-03-15)\n\n"
        f"Send a chart photo to get your position size!"
    )
    bot.send_message(chat_id, text, parse_mode='HTML')

@bot.message_handler(commands=['settings'])
def handle_settings(message):
    chat_id = message.chat.id
    s = get_settings(chat_id)
    cur = s['currency']
    risk_amt = s['capital'] * s['risk'] / 100
    text = (
        f"<b>Your current settings:</b>\n"
        f"💰 Account size: <b>{cur} {s['capital']:,}</b>\n"
        f"⚡ Leverage: <b>1:{s['leverage']}</b>\n"
        f"🎯 Risk per trade: <b>{s['risk']}%</b> ({cur} {risk_amt:,.2f})\n"
        f"💱 Account currency: <b>{cur}</b>"
    )
    bot.send_message(chat_id, text, parse_mode='HTML')

@bot.message_handler(commands=['currency'])
def handle_currency(message):
    chat_id = message.chat.id
    try:
        value = message.text.split()[1].upper()
        if len(value) != 3 or not value.isalpha():
            raise ValueError
        get_settings(chat_id)['currency'] = value
        bot.send_message(chat_id, f"✅ Account currency set to <b>{value}</b>\n\nAll position sizes will now be calculated using {value} as your account base.", parse_mode='HTML')
    except (IndexError, ValueError):
        bot.send_message(chat_id, "❌ Usage: /currency [code] — e.g. /currency GBP or /currency USD")

@bot.message_handler(commands=['capital'])
def handle_capital(message):
    chat_id = message.chat.id
    try:
        value = float(message.text.split()[1])
        if value <= 0:
            raise ValueError
        get_settings(chat_id)['capital'] = value
        bot.send_message(chat_id, f"✅ Account size set to <b>${value:,.2f}</b>", parse_mode='HTML')
    except (IndexError, ValueError):
        bot.send_message(chat_id, "❌ Usage: /capital [amount] — e.g. /capital 30000")

@bot.message_handler(commands=['risk'])
def handle_risk(message):
    chat_id = message.chat.id
    try:
        value = float(message.text.split()[1])
        if value <= 0 or value > 100:
            raise ValueError
        s = get_settings(chat_id)
        s['risk'] = value
        risk_usd = s['capital'] * value / 100
        bot.send_message(chat_id, f"✅ Risk per trade set to <b>{value}%</b> (${risk_usd:,.2f})", parse_mode='HTML')
    except (IndexError, ValueError):
        bot.send_message(chat_id, "❌ Usage: /risk [percent] — e.g. /risk 2")

@bot.message_handler(commands=['leverage'])
def handle_leverage(message):
    chat_id = message.chat.id
    try:
        value = int(message.text.split()[1])
        if value <= 0:
            raise ValueError
        get_settings(chat_id)['leverage'] = value
        bot.send_message(chat_id, f"✅ Leverage set to <b>1:{value}</b>", parse_mode='HTML')
    except (IndexError, ValueError):
        bot.send_message(chat_id, "❌ Usage: /leverage [value] — e.g. /leverage 50")

@bot.message_handler(commands=['journal'])
def handle_journal(message):
    chat_id = message.chat.id
    entries = load_journal()
    user_entries = [e for e in entries if e['chat_id'] == chat_id]

    if not user_entries:
        bot.send_message(chat_id, "📒 Your trading journal is empty. Take a trade first!")
        return

    recent = user_entries[-5:][::-1]
    bot.send_message(chat_id, f"<b>📒 Trading Journal — last {len(recent)} trades:</b>", parse_mode='HTML')

    for entry in recent:
        outcome = entry.get('outcome')
        notes = entry.get('notes')
        uid = entry.get('uid', '')
        outcome_icon = {'Win': '🟢', 'Loss': '🔴', 'Breakeven': '🟡'}.get(outcome, '⏳')

        caption = (
            f"{outcome_icon} <b>{entry['date']}</b>"
            + (f" — <b>{outcome}</b>" if outcome else " — <i>Outcome pending</i>")
            + f"\n\n{entry['analysis']}"
            + (f"\n\n💰 Profit: <b>{notes} USD</b>" if notes else "")
        )

        # Only show button if outcome has NOT been set yet
        markup = None
        if not outcome:
            markup = InlineKeyboardMarkup()
            markup.add(InlineKeyboardButton("✏️ Set outcome", callback_data=f"update:{uid}"))
        else:
            caption += "\n\n🔒 <i>Outcome locked</i>"

        photo_id = entry.get('photo_file_id')
        if photo_id:
            bot.send_photo(chat_id, photo_id, caption=caption, parse_mode='HTML', reply_markup=markup)
        else:
            bot.send_message(chat_id, caption, parse_mode='HTML', reply_markup=markup)

@bot.message_handler(commands=['export'])
def handle_export(message):
    chat_id = message.chat.id
    parts = message.text.split()
    target_date = parts[1] if len(parts) > 1 else time.strftime('%Y-%m-%d')

    entries = load_journal()
    day_entries = [
        e for e in entries
        if e['chat_id'] == chat_id and e.get('date', '').startswith(target_date)
    ]

    if not day_entries:
        bot.send_message(chat_id, f"📭 No trades found for <b>{target_date}</b>.", parse_mode='HTML')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = target_date

    # Header styling
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    headers = ['Date', 'Pair', 'Entry', 'SL', 'Distance', 'LOT', 'Margin', 'Outcome', 'Profit (USD)']
    col_widths = [18, 10, 12, 12, 14, 8, 14, 12, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = width

    outcome_colors = {'Win': 'C6EFCE', 'Loss': 'FFC7CE', 'Breakeven': 'FFEB9C'}
    profit_col = len(headers)  # last column index (1-based)

    for row, entry in enumerate(day_entries, 2):
        f = parse_analysis(entry.get('analysis', ''))
        outcome = entry.get('outcome') or ''
        raw_profit = entry.get('notes') or ''
        try:
            profit_val = float(raw_profit)
        except (ValueError, TypeError):
            profit_val = raw_profit

        row_data = [
            entry.get('date', ''),
            f.get('pair', ''),
            f.get('entry', ''),
            f.get('sl', ''),
            f.get('distance', ''),
            f.get('lot', ''),
            f.get('margin', ''),
            outcome,
            profit_val,
        ]
        outcome_col_index = headers.index('Outcome') + 1
        profit_col_index = headers.index('Profit (USD)') + 1
        fill_color = outcome_colors.get(outcome)
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            if col == outcome_col_index and fill_color:
                cell.fill = PatternFill(fill_type='solid', fgColor=fill_color)
                cell.font = Font(bold=True)

    # Total row
    total_row = len(day_entries) + 2
    total_label_cell = ws.cell(row=total_row, column=profit_col - 1, value='TOTAL')
    total_label_cell.font = Font(bold=True)
    total_label_cell.alignment = Alignment(horizontal='right')

    profit_col_letter = get_column_letter(profit_col)
    sum_cell = ws.cell(
        row=total_row,
        column=profit_col,
        value=f"=SUM({profit_col_letter}2:{profit_col_letter}{total_row - 1})"
    )
    sum_cell.font = Font(bold=True)
    sum_cell.alignment = Alignment(horizontal='center')
    sum_cell.fill = PatternFill(fill_type='solid', fgColor='D9E1F2')

    ws.freeze_panes = 'A2'

    filename = f'journal_{target_date}.xlsx'
    wb.save(filename)

    with open(filename, 'rb') as f:
        bot.send_document(
            chat_id,
            f,
            caption=f"📊 <b>Trading Journal — {target_date}</b>\n{len(day_entries)} trade(s)",
            parse_mode='HTML'
        )
    os.remove(filename)

# ── Chart analysis ────────────────────────────────────────────────────────────

def analyze_chart(image_data, capital, risk, leverage, currency='USD'):
    img = PIL.Image.open(io.BytesIO(image_data))
    risk_amount = capital * risk / 100
    if currency == 'USD':
        currency_note = f"Account currency is USD. Risk amount = {risk_amount:.2f} USD. Use this directly in all lot size formulas."
    else:
        currency_note = (
            f"Account currency is {currency}. Risk amount = {risk_amount:.2f} {currency}.\n"
            f"IMPORTANT: All pip/point values are in USD. Before calculating lots, convert the risk amount to USD:\n"
            f"  Risk in USD = {risk_amount:.2f} {currency} × ({currency}USD exchange rate)\n"
            f"  Use the {currency}USD rate you know or estimate from the chart context.\n"
            f"  Then use Risk_USD in all lot size formulas below."
        )
    prompt = f"""
You are a forex position size calculator. Analyze this trading chart.

CHART RULES:
- Purple line = Stop Loss price
- Teal or Black line = Entry price

ACCOUNT:
- Capital: {capital} {currency}
- Risk: {risk}% = {risk_amount:.2f} {currency} per trade
- Leverage: 1:{leverage}
- {currency_note}

STEP 1 — Identify the instrument and read Entry and SL prices from the chart.

STEP 2 — Determine trade direction:
- If SL < Entry → LONG trade. TP targets are ABOVE entry.
- If SL > Entry → SHORT trade. TP targets are BELOW entry.

STEP 3 — Calculate SL distance and LOT SIZE using the correct formula for the instrument:

  A) Standard Forex pairs (EURUSD, GBPUSD, AUDUSD, NZDUSD, EURGBP, EURAUD, etc.):
     - Pip = 0.0001 | Pip value = $10 per lot
     - SL distance in pips = |Entry - SL| / 0.0001
     - Lots = {risk_amount} / (SL pips × 10)

  B) JPY pairs (USDJPY, EURJPY, GBPJPY, CADJPY, etc.):
     - Pip = 0.01 | Pip value = (100,000 × 0.01) / Entry rate in USD
     - SL distance in pips = |Entry - SL| / 0.01
     - Pip value per lot ≈ 1000 / Entry price (in USD)
     - Lots = {risk_amount} / (SL pips × pip value per lot)

  C) XAUUSD (Gold):
     - Point = $0.01 | Contract = 100 oz | Point value = $1 per lot
     - SL distance in points = |Entry - SL| / 0.01
     - Lots = {risk_amount} / (SL points × 1)

  D) XAGUSD (Silver):
     - Point = $0.001 | Contract = 5,000 oz | Point value = $5 per lot
     - SL distance in points = |Entry - SL| / 0.001
     - Lots = {risk_amount} / (SL points × 5)

  E) Oil (USOIL, UKOIL, WTI, BRENT, CL):
     - Point = $0.01 | Contract = 1,000 barrels | Point value = $10 per lot
